"""
Enrich dmn_questions_export CSV with:
  - sub_specification  → full text from "Sub-Specifications" column (from Excel)
  - accepted_evidence  → English-only lines from "Acceptance Evidence" column (from Excel)

Match key: CSV.dmn_question_text  ↔  Excel "Compliance Questions" (English portion)

Overwrites the CSV in-place.
"""

import csv
import re
import unicodedata
import openpyxl
from pathlib import Path

BASE = Path(__file__).parent
CSV_IN  = BASE / "dmn_questions_export-after-fida-confirmation-23rdFeb.csv"
CSV_OUT = BASE / "dmn_questions_export-after-fida-confirmation-23rdFeb.csv"   # overwrite in-place
XLSX    = BASE / "Compliance Questionnaire_DI Inputs.xlsx"

# ── Sheets that contain actual question data ──────────────────────────────────
DATA_SHEETS = [
    "Data Mgmt Strategy & Governance",
    "Data Quality Management",
    "Data Architecture & Modeling",
    "Master & Ref. Data Management",
    "Data Sec., Priv. & Other Reg.",
    "Data Catalog & Metadata Mgmt.",
    "Data Storage & Operations",
    "Statistics & Analytics",
    "Data Culture & Literacy",
    "Document & Content Management",
    "Data Monetization",
    "Data Sharing,Integ.&Interop.",
]

# ── Helpers ───────────────────────────────────────────────────────────────────
def is_arabic(text: str) -> bool:
    """Return True if the line contains predominantly Arabic script."""
    arabic = sum(1 for c in text if "\u0600" <= c <= "\u06FF")
    return arabic > 2


def english_only(text: str) -> str:
    """Strip Arabic lines from a mixed English/Arabic block."""
    if not text:
        return ""
    lines = text.replace("\r\n", "\n").split("\n")
    english_lines = [l.strip() for l in lines if l.strip() and not is_arabic(l)]
    return "\n".join(english_lines)


def normalize(text: str) -> str:
    """Collapse whitespace, lowercase, and stem common verb tenses for fuzzy matching."""
    if not text:
        return ""
    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r"\s+", " ", text).strip().lower()
    # Normalize common past→present tense variations so CSV & Excel match
    text = re.sub(r"\bestablished\b", "establish", text)
    text = re.sub(r"\bimplemented\b", "implement", text)
    text = re.sub(r"\bdeveloped\b", "develop", text)
    text = re.sub(r"\bconducted\b", "conduct", text)
    text = re.sub(r"\bdefined\b", "define", text)
    text = re.sub(r"\bdesigned\b", "design", text)
    text = re.sub(r"\bidentified\b", "identify", text)
    return text


# ── Load Excel → build lookup: normalized_question_text → (spec_id+name, evidence) ──
def load_excel(path: Path) -> dict:
    print(f"Loading Excel: {path.name}")
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    lookup: dict[str, dict] = {}   # norm_q_text → {sub_spec, accepted_evidence}

    for sheet_name in DATA_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠ Sheet not found: {sheet_name}")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Find header row (has "Compliance Questions")
        header_row_idx = None
        for i, row in enumerate(rows):
            if any(isinstance(c, str) and "Compliance Questions" in c for c in row):
                header_row_idx = i
                break

        if header_row_idx is None:
            print(f"  ⚠ No header row in sheet: {sheet_name}")
            continue

        headers = [str(c).strip() if c else "" for c in rows[header_row_idx]]

        def col(name):
            for j, h in enumerate(headers):
                if name.lower() in h.lower():
                    return j
            return None

        ci_spec_num  = col("Specification Number")
        ci_question  = col("Compliance Questions")
        ci_evidence  = col("Acceptance Evidence")
        ci_sub_spec  = col("Sub-Specifications")

        if ci_question is None:
            print(f"  ⚠ Missing 'Compliance Questions' col in: {sheet_name}")
            continue

        data_rows = rows[header_row_idx + 1:]
        matched = 0
        for row in data_rows:
            def cell(i):
                if i is None or i >= len(row):
                    return ""
                v = row[i]
                return str(v).strip() if v is not None else ""

            raw_q = cell(ci_question)
            if not raw_q:
                continue

            # Extract English portion of the compliance question for matching
            eng_q = english_only(raw_q)
            if not eng_q:
                eng_q = raw_q  # fallback: use full text

            norm_q = normalize(eng_q)
            if not norm_q:
                continue

            # Specification number (e.g. DAM.1.1.1)
            spec_number = cell(ci_spec_num)

            # Sub-specification: full text body
            sub_spec_text = cell(ci_sub_spec)

            # Accepted evidence: English-only lines
            evidence = english_only(cell(ci_evidence))

            lookup[norm_q] = {
                "specification_number": spec_number,
                "sub_specification":    sub_spec_text,
                "accepted_evidence":    evidence,
            }
            matched += 1

        print(f"  ✓ {sheet_name}: {matched} rows indexed")

    print(f"Total Excel entries indexed: {len(lookup)}")
    return lookup


# ── Read CSV, enrich, write back ──────────────────────────────────────────────
def enrich_csv(csv_path: Path, out_path: Path, lookup: dict):
    print(f"\nReading CSV: {csv_path.name}")
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        orig_fields = reader.fieldnames or []
        rows = list(reader)

    print(f"  Rows: {len(rows)}, Columns: {len(orig_fields)}")

    # Add new columns if not already present
    new_fields = list(orig_fields)
    if "specification_number" not in new_fields:
        new_fields.append("specification_number")
        print("  + Added column: specification_number")
    if "sub_specification" not in new_fields:
        new_fields.append("sub_specification")
        print("  + Added column: sub_specification")
    if "accepted_evidence" not in new_fields:
        new_fields.append("accepted_evidence")
        print("  + Added column: accepted_evidence")

    matched = 0
    unmatched = []

    for row in rows:
        q_text = row.get("dmn_question_text", "")
        norm_q = normalize(english_only(q_text) or q_text)

        hit = lookup.get(norm_q)
        if hit:
            row["specification_number"] = hit["specification_number"]
            row["sub_specification"]    = hit["sub_specification"]
            row["accepted_evidence"]    = hit["accepted_evidence"]
            matched += 1
        else:
            # Try partial match: first 100 chars
            short = norm_q[:100]
            hit = next((v for k, v in lookup.items() if k.startswith(short)), None)
            if hit:
                row["specification_number"] = hit["specification_number"]
                row["sub_specification"]    = hit["sub_specification"]
                row["accepted_evidence"]    = hit["accepted_evidence"]
                matched += 1
            else:
                row.setdefault("specification_number", "")
                row.setdefault("sub_specification", "")
                row.setdefault("accepted_evidence", "")
                unmatched.append(q_text[:80])

    print(f"\nMatched: {matched}/{len(rows)}")
    if unmatched:
        print(f"Unmatched ({len(unmatched)}):")
        for u in unmatched[:10]:
            print(f"  - {u}")
        if len(unmatched) > 10:
            print(f"  ... and {len(unmatched)-10} more")

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=new_fields)
        writer.writeheader()
        writer.writerows(rows)

    print(f"\n✅ Saved: {out_path.name}")


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    lookup = load_excel(XLSX)
    enrich_csv(CSV_IN, CSV_OUT, lookup)
