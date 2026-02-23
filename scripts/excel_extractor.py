#!/usr/bin/env python3
"""
Excel Extractor for n8n Compliance Workflow A
==============================================
Adapted from the Universal Excel Table Extractor.

Called by n8n's Execute Command node:
    python3 /scripts/excel_extractor.py /tmp/n8n_processing/<prefix>input.xlsx <originalFileName>

Outputs a single JSON object to stdout matching Workflow A's extraction contract:
{
  "filePrefix":       "<prefix>",
  "originalFileName": "<name.xlsx>",
  "totalPages":       <int>,       # == number of sheets extracted
  "totalWords":       <int>,
  "hasDiagrams":      false,       # Excel has no image analysis
  "fullDocument":     "<text>",    # All sheets concatenated
  "pages": [
    {
      "pageNumber":     1,
      "sheetName":      "Sheet1",
      "text":           "<CSV-like text>",
      "wordCount":      <int>,
      "rowCount":       <int>,
      "columnCount":    <int>,
      "visionAnalysis": {},
      "isDiagram":      false
    }, ...
  ],
  "metadata": {
    "sheetsExtracted": [...],
    "sheetsSkipped":   [...],
    "extractor":       "excel_extractor.py"
  }
}

Exit codes:
  0 — success (JSON printed to stdout)
  1 — file not found or unsupported format
  2 — extraction error (encrypted/corrupt file)
"""

import os
import sys
import json
import re
from pathlib import Path

try:
    import openpyxl
    import pandas as pd
except ImportError as e:
    print(json.dumps({
        "error": f"Missing dependency: {e}. Run: pip3 install openpyxl pandas",
        "exitCode": 1
    }), flush=True)
    sys.exit(1)


# ── Tuning constants (same as original, proven on real files) ──────────────────
MIN_DENSE_ROWS     = 2    # min rows with ≥ MIN_FILLED_CELLS values to qualify a sheet
MIN_FILLED_CELLS   = 3    # cells per row to count as "dense"
MAX_AVG_HEADER_LEN = 80   # max avg char length per cell in a header row
MIN_UNIQUE_RATIO   = 0.5  # min fraction of unique values in a header row


# ── Helpers ───────────────────────────────────────────────────────────────────

def is_filled(value) -> bool:
    """True only if cell has meaningful text/number content."""
    return value is not None and str(value).strip() != ""


def unmerge_and_fill(ws) -> None:
    """
    Unmerge every merged cell range in-place.
    Each cell in the range gets the value from the top-left cell.
    Ensures merged hierarchical columns become plain repeating values.
    """
    for merge_range in list(ws.merged_cells.ranges):
        top_left_value = ws.cell(merge_range.min_row, merge_range.min_col).value
        ws.unmerge_cells(str(merge_range))
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                ws.cell(row, col).value = top_left_value


def get_matrix(ws) -> list:
    """Return all cell values as a list of lists (rows × cols)."""
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def count_dense_rows(matrix) -> int:
    return sum(
        1 for row in matrix
        if sum(1 for v in row if is_filled(v)) >= MIN_FILLED_CELLS
    )


def find_header_row(matrix) -> int | None:
    """
    Find the best header row by scoring every candidate row.
    Score = unique_filled_count × dense_rows_below_it
    """
    best_idx, best_score = None, 0

    for i, row in enumerate(matrix):
        filled = [v for v in row if is_filled(v)]
        if len(filled) < MIN_FILLED_CELLS:
            continue

        avg_len = sum(len(str(v)) for v in filled) / len(filled)
        if avg_len >= MAX_AVG_HEADER_LEN:
            continue

        unique_vals = set(str(v).strip() for v in filled)
        if len(unique_vals) / len(filled) < MIN_UNIQUE_RATIO:
            continue

        dense_below = sum(
            1 for j in range(i + 1, len(matrix))
            if sum(1 for v in matrix[j] if is_filled(v)) >= MIN_FILLED_CELLS
        )
        score = len(unique_vals) * dense_below

        if score > best_score:
            best_score = score
            best_idx = i

    return best_idx if best_score > 0 else None


def build_dataframe(matrix, header_row_idx: int) -> pd.DataFrame:
    """Build a clean DataFrame with deduplicated columns, dropping empty rows/cols."""
    header = matrix[header_row_idx]

    seen = {}
    columns = []
    for col in header:
        name = str(col).strip() if is_filled(col) else "Unnamed"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        columns.append(name)

    n_cols = len(columns)
    rows = [
        (list(row) + [None] * n_cols)[:n_cols]
        for row in matrix[header_row_idx + 1:]
    ]

    df = pd.DataFrame(rows, columns=columns)
    df = df.dropna(how="all")
    df = df[~df.apply(lambda r: all(not is_filled(v) for v in r), axis=1)]
    df = df.dropna(axis=1, how="all")
    empty_unnamed = [
        c for c in df.columns
        if c.startswith("Unnamed") and df[c].isna().all()
    ]
    df = df.drop(columns=empty_unnamed, errors="ignore")
    df = df.map(lambda v: str(v).strip() if is_filled(v) else "")

    return df.reset_index(drop=True)


def df_to_text(df: pd.DataFrame, sheet_name: str) -> str:
    """
    Convert a DataFrame to readable plain text for the LLM.
    Format: Sheet name header, then pipe-separated rows.
    This preserves column relationships better than raw CSV for RAG.
    """
    lines = [f"=== Sheet: {sheet_name} ==="]
    lines.append(" | ".join(df.columns.tolist()))
    lines.append("-" * 80)
    for _, row in df.iterrows():
        lines.append(" | ".join(str(v) if v else "" for v in row))
    return "\n".join(lines)


def word_count(text: str) -> int:
    return len(text.split()) if text.strip() else 0


# ── Main extractor ─────────────────────────────────────────────────────────────

def extract(file_path: str, original_file_name: str) -> dict:
    fp = Path(file_path)

    # Derive filePrefix from the temp filename (strip "input.xlsx" suffix)
    file_prefix = fp.name  # e.g. "1708675200000_abc1_input.xlsx"
    match = re.match(r'^(.+?)input\.[^.]+$', fp.name)
    if match:
        file_prefix = match.group(1)

    try:
        wb = openpyxl.load_workbook(str(fp), data_only=True)
    except Exception as e:
        err_msg = str(e)
        if "encrypted" in err_msg.lower() or "password" in err_msg.lower():
            print(json.dumps({
                "error": "Excel file is password-protected and cannot be processed.",
                "errorCode": "EXCEL_ENCRYPTED",
                "originalFileName": original_file_name
            }), flush=True)
        else:
            print(json.dumps({
                "error": f"Failed to open Excel file: {err_msg}",
                "errorCode": "EXCEL_CORRUPT",
                "originalFileName": original_file_name
            }), flush=True)
        sys.exit(2)

    pages = []
    sheets_extracted = []
    sheets_skipped = []
    page_number = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        unmerge_and_fill(ws)
        matrix = get_matrix(ws)

        dense = count_dense_rows(matrix)
        if dense < MIN_DENSE_ROWS:
            sheets_skipped.append(sheet_name)
            continue

        header_idx = find_header_row(matrix)
        if header_idx is None:
            sheets_skipped.append(sheet_name)
            continue

        df = build_dataframe(matrix, header_idx)
        if df.empty:
            sheets_skipped.append(sheet_name)
            continue

        page_number += 1
        text = df_to_text(df, sheet_name)
        wc = word_count(text)

        sheets_extracted.append(sheet_name)
        pages.append({
            "pageNumber":    page_number,
            "sheetName":     sheet_name,
            "text":          text,
            "wordCount":     wc,
            "rowCount":      len(df),
            "columnCount":   len(df.columns),
            "visionAnalysis": {},   # No vision for spreadsheets
            "isDiagram":     False
        })

    if not pages:
        # No extractable sheets — return graceful empty result, not an error
        print(json.dumps({
            "filePrefix":       file_prefix,
            "originalFileName": original_file_name,
            "totalPages":       0,
            "totalWords":       0,
            "hasDiagrams":      False,
            "fullDocument":     "",
            "pages":            [],
            "metadata": {
                "sheetsExtracted": [],
                "sheetsSkipped":   sheets_skipped,
                "extractor":       "excel_extractor.py",
                "note":            "No data sheets found in this Excel file."
            }
        }), flush=True)
        sys.exit(0)

    full_document = "\n\n".join(p["text"] for p in pages)
    total_words = sum(p["wordCount"] for p in pages)

    result = {
        "filePrefix":       file_prefix,
        "originalFileName": original_file_name,
        "totalPages":       len(pages),
        "totalWords":       total_words,
        "hasDiagrams":      False,
        "fullDocument":     full_document,
        "pages":            pages,
        "metadata": {
            "sheetsExtracted": sheets_extracted,
            "sheetsSkipped":   sheets_skipped,
            "extractor":       "excel_extractor.py"
        }
    }

    print(json.dumps(result, ensure_ascii=False, default=str), flush=True)
    sys.exit(0)


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({
            "error": "Usage: excel_extractor.py <file_path> [original_file_name]",
            "errorCode": "MISSING_ARGS"
        }), flush=True)
        sys.exit(1)

    file_path = sys.argv[1]
    original_name = sys.argv[2] if len(sys.argv) > 2 else Path(file_path).name

    if not os.path.exists(file_path):
        print(json.dumps({
            "error": f"File not found: {file_path}",
            "errorCode": "FILE_NOT_FOUND",
            "originalFileName": original_name
        }), flush=True)
        sys.exit(1)

    extract(file_path, original_name)
