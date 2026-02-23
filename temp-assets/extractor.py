#!/usr/bin/env python3
"""
Universal Excel Table Extractor
================================
Throws any .xlsx at it → finds real data tables → extracts them.
No template assumptions. Works on any Excel file.

Usage:
    python extractor.py --file path/to/file.xlsx
    python extractor.py --file path/to/file.xlsx --output ./my_output
    python extractor.py --folder ./excel_files --output ./output

Output per file:
    output/<filename>__<SheetName>.csv  — one CSV per data sheet found
    output/<filename>__all_sheets.json  — all sheets combined
"""

import os
import json
import argparse
import pandas as pd
import openpyxl
from pathlib import Path

# ── Tuning constants ───────────────────────────────────────────────────────────
MIN_DENSE_ROWS     = 2   # min rows with ≥ MIN_FILLED_CELLS values to qualify a sheet
MIN_FILLED_CELLS   = 3   # cells per row to count as "dense"
MAX_AVG_HEADER_LEN = 80  # max avg char length per cell in a header row (labels ≠ paragraphs)
MIN_UNIQUE_RATIO   = 0.5 # min fraction of unique values in a header row


# ── Helpers ───────────────────────────────────────────────────────────────────

def is_filled(value) -> bool:
    """True only if cell has meaningful text/number content."""
    return value is not None and str(value).strip() != ""


def unmerge_and_fill(ws) -> None:
    """
    Unmerge every merged cell range in-place.
    Each cell in the range gets the value from the top-left cell.
    This ensures merged hierarchical columns (e.g. Control Number spanning
    multiple rows) become plain repeating values that pandas can read.
    """
    for merge_range in list(ws.merged_cells.ranges):
        top_left_value = ws.cell(merge_range.min_row, merge_range.min_col).value
        ws.unmerge_cells(str(merge_range))
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                ws.cell(row, col).value = top_left_value


def get_matrix(ws) -> list:
    """Return all cell values as a list of lists (rows × cols)."""
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def count_dense_rows(matrix) -> int:
    """Count rows that have at least MIN_FILLED_CELLS non-empty cells."""
    return sum(
        1 for row in matrix
        if sum(1 for v in row if is_filled(v)) >= MIN_FILLED_CELLS
    )


def find_header_row(matrix) -> int | None:
    """
    Find the best header row by scoring every candidate row.

    Score = unique_filled_count × dense_rows_below_it

    Why this works:
    - unique_filled_count: real headers have distinct column labels.
      A merged title row (e.g. "Analytics Use Case Library" spanning 10 cols)
      fails this because all repeated cells are identical after unmerging.
    - dense_rows_below: the true header sits above the most data rows.
      Title blocks near the top have little real data below them.
    - avg_len guard: skips paragraph/description rows (long sentences ≠ labels).

    Returns the row index with the highest score, or None if nothing qualifies.
    """
    best_idx, best_score = None, 0

    for i, row in enumerate(matrix):
        filled = [v for v in row if is_filled(v)]
        if len(filled) < MIN_FILLED_CELLS:
            continue

        # Guard: skip paragraph rows
        avg_len = sum(len(str(v)) for v in filled) / len(filled)
        if avg_len >= MAX_AVG_HEADER_LEN:
            continue

        # Guard: skip rows where most values are identical (unmerged title rows)
        unique_vals = set(str(v).strip() for v in filled)
        if len(unique_vals) / len(filled) < MIN_UNIQUE_RATIO:
            continue

        # Score: how many dense rows come after this candidate?
        dense_below = sum(
            1 for j in range(i + 1, len(matrix))
            if sum(1 for v in matrix[j] if is_filled(v)) >= MIN_FILLED_CELLS
        )
        score = len(unique_vals) * dense_below

        if score > best_score:
            best_score = score
            best_idx = i

    return best_idx if best_score > 0 else None


def build_dataframe(matrix, header_row_idx: int) -> pd.DataFrame:
    """
    Build a clean DataFrame from the matrix using header_row_idx as columns.

    - Deduplicates column names (appends _1, _2 for duplicates)
    - Pads/trims rows to match column count
    - Drops fully-empty rows
    - Drops columns that are entirely empty
    - Strips whitespace from all string values
    """
    header = matrix[header_row_idx]

    # Build clean, deduplicated column names
    seen = {}
    columns = []
    for col in header:
        name = str(col).strip() if is_filled(col) else "Unnamed"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        columns.append(name)

    n_cols = len(columns)
    rows = [
        (list(row) + [None] * n_cols)[:n_cols]
        for row in matrix[header_row_idx + 1:]
    ]

    df = pd.DataFrame(rows, columns=columns)

    # Drop fully empty rows
    df = df.dropna(how="all")
    df = df[~df.apply(lambda r: all(not is_filled(v) for v in r), axis=1)]

    # Drop fully empty columns (incl. trailing Unnamed cols from wide sheets)
    df = df.dropna(axis=1, how="all")
    empty_unnamed = [
        c for c in df.columns
        if c.startswith("Unnamed") and df[c].isna().all()
    ]
    df = df.drop(columns=empty_unnamed, errors="ignore")

    # Normalise cell values
    df = df.map(lambda v: str(v).strip() if is_filled(v) else None)

    return df.reset_index(drop=True)


# ── Core processor ─────────────────────────────────────────────────────────────

def process_file(file_path: str, output_dir: str) -> dict:
    """
    Process a single .xlsx file.
    Returns a result dict and writes CSV + JSON to output_dir.
    """
    file_path  = Path(file_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"  FILE : {file_path.name}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(file_path, data_only=True)

    result = {
        "filename"         : file_path.name,
        "sheets_extracted" : [],
        "sheets_skipped"   : [],
        "data"             : {}
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Step 1 — unmerge all cells
        unmerge_and_fill(ws)

        # Step 2 — raw matrix
        matrix = get_matrix(ws)

        # Step 3 — qualify sheet
        dense = count_dense_rows(matrix)
        if dense < MIN_DENSE_ROWS:
            print(f"  ⏭  SKIP  '{sheet_name}'  (only {dense} dense rows)")
            result["sheets_skipped"].append(sheet_name)
            continue

        # Step 4 — find header
        header_idx = find_header_row(matrix)
        if header_idx is None:
            print(f"  ⏭  SKIP  '{sheet_name}'  (no header row detected)")
            result["sheets_skipped"].append(sheet_name)
            continue

        # Step 5 — extract table
        df = build_dataframe(matrix, header_idx)
        if df.empty:
            print(f"  ⏭  SKIP  '{sheet_name}'  (table is empty after extraction)")
            result["sheets_skipped"].append(sheet_name)
            continue

        # ── success ──
        print(f"  ✅ EXTRACTED  '{sheet_name}'")
        print(f"     Header row : {header_idx}")
        print(f"     Shape      : {len(df)} rows × {len(df.columns)} cols")
        print(f"     Columns    : {list(df.columns)}")

        # Save individual CSV
        safe_name = "".join(
            c if c.isalnum() or c in "._- " else "_"
            for c in sheet_name
        ).strip()
        csv_path = output_dir / f"{file_path.stem}__{safe_name}.csv"
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        print(f"     Saved  →   {csv_path.name}")

        result["sheets_extracted"].append(sheet_name)
        result["data"][sheet_name] = df.to_dict(orient="records")

    # Save combined JSON
    json_path = output_dir / f"{file_path.stem}__all_sheets.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2, default=str)

    print(f"\n  📦 JSON  →  {json_path.name}")
    print(f"  ✔  {len(result['sheets_extracted'])} extracted  |  "
          f"{len(result['sheets_skipped'])} skipped\n")

    return result


# ── CLI ────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Universal Excel Table Extractor — no template config needed"
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--file",   help="Path to a single .xlsx file")
    group.add_argument("--folder", help="Path to a folder; all .xlsx files are processed")
    parser.add_argument("--output", default="./output",
                        help="Output directory (default: ./output)")
    args = parser.parse_args()

    if args.file:
        if not os.path.exists(args.file):
            print(f"Error: File not found → {args.file}")
            return
        process_file(args.file, args.output)

    elif args.folder:
        xlsx_files = list(Path(args.folder).glob("*.xlsx"))
        if not xlsx_files:
            print(f"No .xlsx files found in {args.folder}")
            return
        print(f"Found {len(xlsx_files)} file(s) in {args.folder}")
        for f in xlsx_files:
            process_file(str(f), args.output)


if __name__ == "__main__":
    main()
